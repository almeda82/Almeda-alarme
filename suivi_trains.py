import time
import datetime
import os
from PIL import ImageGrab
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ------------------------------------------------------------------------------
# CONFIGURATION ET GESTION DU FICHIER EXCEL (.XLSX)
# ------------------------------------------------------------------------------
FICHIER_EXCEL = "registre_circulations_blida.xlsx"

def initialiser_excel():
    # Vérification purge 15 jours
    if os.path.exists(FICHIER_EXCEL):
        t_creation = os.path.getctime(FICHIER_EXCEL)
        age_jours = (time.time() - t_creation) / (24 * 3600)
        if age_jours >= 15:
            os.remove(FICHIER_EXCEL)
            print("[PURGE] Fichier réinitialisé (limite de 15 jours atteinte).")

    if not os.path.exists(FICHIER_EXCEL):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Circulations"
        
        # En-tête Date à 00h00
        date_jour = datetime.date.today().strftime('%d/%m/%Y')
        ws.merge_cells('A1:R1')
        ws['A1'] = f"REGISTRE DU {date_jour} - GARE DE BLIDA"
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="1F4E78", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal="center")

        # Titres Voies
        ws.merge_cells('A2:I2')
        ws['A2'] = "VOIE 1 (IMPAIR - Sens Alger -> El Affroun)"
        ws['A2'].font = Font(bold=True, color="FFFFFF")
        ws['A2'].fill = PatternFill(start_color="203764", fill_type="solid")
        ws['A2'].alignment = Alignment(horizontal="center")

        ws.merge_cells('K2:R2')
        ws['K2'] = "VOIE 2 (PAIR - Sens El Affroun -> Alger)"
        ws['K2'].font = Font(bold=True, color="FFFFFF")
        ws['K2'].fill = PatternFill(start_color="203764", fill_type="solid")
        ws['K2'].alignment = Alignment(horizontal="center")

        # Colonnes V1
        cols_v1 = ["N°", "C1", "Z3", "Z5", "Z7", "C15", "Z11", "Z17", "Z13"]
        for idx, col in enumerate(cols_v1, start=1):
            cell = ws.cell(row=3, column=idx, value=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Colonnes V2
        cols_v2 = ["N°", "C24", "Z36", "Z34", "Z10", "C8", "Z8", "Z6", "Z4"]
        for idx, col in enumerate(cols_v2, start=11):
            cell = ws.cell(row=3, column=idx, value=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        wb.save(FICHIER_EXCEL)

def ajouter_train_excel(voie, num_train, donnees_occ, donnees_lib):
    wb = openpyxl.load_workbook(FICHIER_EXCEL)
    ws = wb.active

    # Trouver la première ligne vide selon la voie
    col_dep = 1 if voie == "V1" else 11
    row = 4
    while ws.cell(row=row, column=col_dep).value is not None:
        row += 2  # Chaque train prend 2 lignes

    # Fusion des 2 lignes pour le numéro de train (Tr1, Tr2...)
    ws.merge_cells(start_row=row, start_column=col_dep, end_row=row+1, end_column=col_dep)
    cell_num = ws.cell(row=row, column=col_dep, value=f"Tr {num_train}")
    cell_num.alignment = Alignment(horizontal="center", vertical="center")
    cell_num.font = Font(bold=True)

    font_rouge = Font(color="FF0000", bold=True)
    font_noir = Font(color="000000")

    # Inscription Ligne Haut (Occupation - Rouge)
    for idx, val in enumerate(donnees_occ, start=col_dep+1):
        cell = ws.cell(row=row, column=idx, value=val)
        cell.font = font_rouge
        cell.alignment = Alignment(horizontal="center")

    # Inscription Ligne Bas (Libération - Noir)
    for idx, val in enumerate(donnees_lib, start=col_dep+1):
        cell = ws.cell(row=row+1, column=idx, value=val)
        cell.font = font_noir
        cell.alignment = Alignment(horizontal="center")

    wb.save(FICHIER_EXCEL)
    print(f"[{datetime.datetime.now().strftime('%H:%M:%S')}] Train Tr {num_train} ({voie}) inscrit dans le fichier Excel.")

# ------------------------------------------------------------------------------
# DÉTECTION ET LECTURE TCO
# ------------------------------------------------------------------------------
def charger_config(fichier="config.txt"):
    coords = {}
    if not os.path.exists(fichier):
        return None
    with open(fichier, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#"):
                cle, valeur = line.split("=")
                coords[cle.strip()] = int(valeur.strip())
    return coords

def lire_pixel(x, y):
    img = ImageGrab.grab(bbox=(x, y, x + 1, y + 1))
    return img.getpixel((0, 0))

def est_rouge(rgb):
    return rgb[0] > 150 and rgb[1] < 70 and rgb[2] < 70

def est_vert(rgb):
    return rgb[1] > 150 and rgb[0] < 100 and rgb[2] < 100

def format_h():
    return datetime.datetime.now().strftime('%H:%M:%S')

# ------------------------------------------------------------------------------
# SURVEILLANCE
# ------------------------------------------------------------------------------
def surveiller():
    cfg = charger_config()
    if not cfg:
        print("Erreur : fichier config.txt introuvable !")
        return

    initialiser_excel()

    tr_pair_count = 1
    tr_impair_count = 1

    etat_p = "ATTENTE"
    etat_i = "ATTENTE"

    occ_p, lib_p = {}, {}
    occ_i, lib_i = {}, {}

    print("Surveillance TCO lancée avec format Excel optimisé...")

    while True:
        # --- TRAIN PAIR (VOIE 2) ---
        c_c24 = lire_pixel(cfg['C24_X'], cfg['C24_Y'])
        c_z36 = lire_pixel(cfg['Z36_X'], cfg['Z36_Y'])
        c_z34 = lire_pixel(cfg['Z34_X'], cfg['Z34_Y'])
        c_z10 = lire_pixel(cfg['Z10_X'], cfg['Z10_Y'])
        c_c8  = lire_pixel(cfg['C8_X'],  cfg['C8_Y'])
        c_z8  = lire_pixel(cfg['Z8_X'],  cfg['Z8_Y'])
        c_z6  = lire_pixel(cfg['Z6_X'],  cfg['Z6_Y'])
        c_z4  = lire_pixel(cfg['Z4_X'],  cfg['Z4_Y'])

        if est_vert(c_c24) and 'C24' not in occ_p:
            occ_p['C24'] = format_h()

        if etat_p == "ATTENTE" and est_rouge(c_z36):
            etat_p = "EN_COURS"
            occ_p['Z36'] = format_h()
            if est_rouge(c_c24): lib_p['C24'] = format_h()

        if etat_p == "EN_COURS":
            if not est_rouge(c_z36) and 'Z36' in occ_p and 'Z36' not in lib_p: lib_p['Z36'] = format_h()
            if est_rouge(c_z34) and 'Z34' not in occ_p: occ_p['Z34'] = format_h()
            if not est_rouge(c_z34) and 'Z34' in occ_p and 'Z34' not in lib_p: lib_p['Z34'] = format_h()
            if est_rouge(c_z10) and 'Z10' not in occ_p: occ_p['Z10'] = format_h()
            if est_vert(c_c8) and 'C8' not in occ_p: occ_p['C8'] = format_h()
            if 'Z10' in occ_p and not est_rouge(c_z10) and 'Z10' not in lib_p:
                lib_p['Z10'] = format_h()
                if est_rouge(c_c8): lib_p['C8'] = format_h()
            if est_rouge(c_z8) and 'Z8' not in occ_p: occ_p['Z8'] = format_h()
            if not est_rouge(c_z8) and 'Z8' in occ_p and 'Z8' not in lib_p: lib_p['Z8'] = format_h()
            if est_rouge(c_z6) and 'Z6' not in occ_p: occ_p['Z6'] = format_h()
            if not est_rouge(c_z6) and 'Z6' in occ_p and 'Z6' not in lib_p: lib_p['Z6'] = format_h()
            if est_rouge(c_z4) and 'Z4' not in occ_p: occ_p['Z4'] = format_h()
            if not est_rouge(c_z4) and 'Z4' in occ_p and 'Z4' not in lib_p:
                lib_p['Z4'] = format_h()
                # Fin de passage train Pair
                d_occ = [occ_p.get(k, '') for k in ['C24','Z36','Z34','Z10','C8','Z8','Z6','Z4']]
                d_lib = [lib_p.get(k, '') for k in ['C24','Z36','Z34','Z10','C8','Z8','Z6','Z4']]
                ajouter_train_excel("V2", tr_pair_count, d_occ, d_lib)
                tr_pair_count += 1
                etat_p = "ATTENTE"
                occ_p, lib_p = {}, {}

        # --- TRAIN IMPAIR (VOIE 1) ---
        c_c1  = lire_pixel(cfg['C1_X'],  cfg['C1_Y'])
        c_z3  = lire_pixel(cfg['Z3_X'],  cfg['Z3_Y'])
        c_z5  = lire_pixel(cfg['Z5_X'],  cfg['Z5_Y'])
        c_z7  = lire_pixel(cfg['Z7_X'],  cfg['Z7_Y'])
        c_c15 = lire_pixel(cfg['C15_X'], cfg['C15_Y'])
        c_z11 = lire_pixel(cfg['Z11_X'], cfg['Z11_Y'])
        c_z17 = lire_pixel(cfg['Z17_X'], cfg['Z17_Y'])
        c_z13 = lire_pixel(cfg['Z13_X'], cfg['Z13_Y'])

        if est_vert(c_c1) and 'C1' not in occ_i:
            occ_i['C1'] = format_h()

        if etat_i == "ATTENTE" and est_rouge(c_z3):
            etat_i = "EN_COURS"
            occ_i['Z3'] = format_h()
            if est_rouge(c_c1): lib_i['C1'] = format_h()

        if etat_i == "EN_COURS":
            if not est_rouge(c_z3) and 'Z3' in occ_i and 'Z3' not in lib_i: lib_i['Z3'] = format_h()
            if est_rouge(c_z5) and 'Z5' not in occ_i: occ_i['Z5'] = format_h()
            if not est_rouge(c_z5) and 'Z5' in occ_i and 'Z5' not in lib_i: lib_i['Z5'] = format_h()
            if est_rouge(c_z7) and 'Z7' not in occ_i: occ_i['Z7'] = format_h()
            if est_vert(c_c15) and 'C15' not in occ_i: occ_i['C15'] = format_h()
            if 'Z7' in occ_i and not est_rouge(c_z7) and 'Z7' not in lib_i:
                lib_i['Z7'] = format_h()
                if est_rouge(c_c15): lib_i['C15'] = format_h()
            if est_rouge(c_z11) and 'Z11' not in occ_i: occ_i['Z11'] = format_h()
            if not est_rouge(c_z11) and 'Z11' in occ_i and 'Z11' not in lib_i: lib_i['Z11'] = format_h()
            if est_rouge(c_z17) and 'Z17' not in occ_i: occ_i['Z17'] = format_h()
            if not est_rouge(c_z17) and 'Z17' in occ_i and 'Z17' not in lib_i: lib_i['Z17'] = format_h()
            if est_rouge(c_z13) and 'Z13' not in occ_i: occ_i['Z13'] = format_h()
            if not est_rouge(c_z13) and 'Z13' in occ_i and 'Z13' not in lib_i:
                lib_i['Z13'] = format_h()
                # Fin de passage train Impair
                d_occ = [occ_i.get(k, '') for k in ['C1','Z3','Z5','Z7','C15','Z11','Z17','Z13']]
                d_lib = [lib_i.get(k, '') for k in ['C1','Z3','Z5','Z7','C15','Z11','Z17','Z13']]
                ajouter_train_excel("V1", tr_impair_count, d_occ, d_lib)
                tr_impair_count += 1
                etat_i = "ATTENTE"
                occ_i, lib_i = {}, {}

        time.sleep(0.3)

if __name__ == "__main__":
    surveiller()